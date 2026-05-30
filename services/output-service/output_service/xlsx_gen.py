"""XLSX output generator (M7-17, M7-18, M7-19).

Generates policy comparison spreadsheets with formatted sheets.
"""

from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_CITATION_CONFIDENCE_FILLS = {
    "direct": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "fuzzy": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    "uncertain": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
}

_THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
_BODY_FONT = Font(name="微软雅黑", size=10)
_BOLD_FONT = Font(name="微软雅黑", size=11, bold=True)


class XlsxGenerator:
    """Generates multi-sheet XLSX output for policy comparisons."""

    def generate(
        self,
        title: str,
        sections: list[dict[str, Any]],
        citations: list[dict[str, Any]],
        metadata: dict[str, Any] | None = None,
    ) -> bytes:
        """生成 an XLSX file with comparison matrix, citation list, and data summary.

        Args:
            title: Report title.
            sections: Analysis sections.
            citations: Citation data.
            metadata: Optional metadata including comparison_matrix.

        Returns:
            XLSX file bytes.
        """
        meta = metadata or {}
        wb = Workbook()

        self._build_comparison_sheet(wb, title, meta)
        self._build_citations_sheet(wb, citations)
        self._build_data_summary_sheet(wb, meta)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _build_comparison_sheet(self, wb: Workbook, title: str, meta: dict[str, Any]) -> None:
        """M7-17: Build the comparison analysis sheet."""
        ws = wb.active
        ws.title = "对比分析"

        matrix = meta.get("comparison_matrix", {})
        policy_names: list[str] = matrix.get("policy_names", [])
        rows_data: list[dict[str, Any]] = matrix.get("rows", [])

        # Title
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1 + len(policy_names))
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = Font(name="微软雅黑", size=16, bold=True, color="1F4E79")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 36

        # Date
        date_str = meta.get("date", "")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=1 + len(policy_names))
        date_cell = ws.cell(row=2, column=1, value=f"生成日期: {date_str}")
        date_cell.font = Font(name="微软雅黑", size=11, color="666666")
        ws.row_dimensions[2].height = 24

        # Headers
        header_row = 4
        ws.cell(row=header_row, column=1, value="比较维度").font = Font(
            name="微软雅黑", size=11, bold=True, color="FFFFFF"
        )
        ws.cell(row=header_row, column=1).fill = _HEADER_FILL
        ws.cell(row=header_row, column=1).border = _THIN_BORDER
        ws.column_dimensions["A"].width = 22

        for i, pname in enumerate(policy_names):
            col = 2 + i
            ws.cell(row=header_row, column=col, value=pname).font = Font(
                name="微软雅黑", size=11, bold=True, color="FFFFFF"
            )
            ws.cell(row=header_row, column=col).fill = _HEADER_FILL
            ws.cell(row=header_row, column=col).border = _THIN_BORDER
            ws.column_dimensions[get_column_letter(col)].width = 35

        ws.row_dimensions[header_row].height = 28

        # Data rows
        for r_idx, row_data in enumerate(rows_data):
            row_num = header_row + 1 + r_idx
            dim_cell = ws.cell(row=row_num, column=1, value=row_data.get("dimension", ""))
            dim_cell.font = _BOLD_FONT
            dim_cell.fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
            dim_cell.alignment = Alignment(horizontal="left", vertical="top")
            dim_cell.border = _THIN_BORDER

            for p_idx, pname in enumerate(policy_names):
                col = 2 + p_idx
                cell_value = row_data.get("values", {}).get(pname, "")
                cell = ws.cell(row=row_num, column=col, value=cell_value)
                cell.font = _BODY_FONT
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                cell.border = _THIN_BORDER

            ws.row_dimensions[row_num].height = 80

    def _build_citations_sheet(self, wb: Workbook, citations: list[dict[str, Any]]) -> None:
        """M7-18: Build the citation list sheet."""
        ws = wb.create_sheet("引用清单")

        headers = ["序号", "来源文档", "作者/发布机构", "页码范围", "引用上下文", "置信度", "置信度说明"]
        col_widths = [6, 30, 20, 12, 45, 12, 16]

        confidence_labels = {
            "direct": "直接引用 — 页码精确匹配",
            "fuzzy": "段落匹配 — 语义相似",
            "uncertain": "模型推理 — 无直接原文依据",
        }

        for c_idx, (header, width) in enumerate(zip(headers, col_widths, strict=True), 1):
            cell = ws.cell(row=1, column=c_idx, value=header)
            cell.font = _HEADER_FONT
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _THIN_BORDER
            ws.column_dimensions[get_column_letter(c_idx)].width = width

        ws.row_dimensions[1].height = 28

        for i, cit in enumerate(citations, 1):
            row = i + 1
            confidence = cit.get("confidence", "")

            values = [
                i,
                cit.get("document_title", cit.get("ref_id", "")),
                cit.get("authors", ""),
                cit.get("source_page", cit.get("page_range", "-")),
                cit.get("sentence", ""),
                confidence,
                confidence_labels.get(confidence, ""),
            ]

            for c_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=c_idx, value=value)
                cell.font = _BODY_FONT
                cell.border = _THIN_BORDER
                if c_idx == 6 and confidence in _CITATION_CONFIDENCE_FILLS:
                    cell.fill = _CITATION_CONFIDENCE_FILLS[confidence]

            ws.row_dimensions[row].height = 24

    def _build_data_summary_sheet(self, wb: Workbook, meta: dict[str, Any]) -> None:
        """M7-19: Build the optional data summary sheet."""
        metrics = meta.get("data_metrics")
        if not metrics:
            return

        ws = wb.create_sheet("数据摘要")

        ws.merge_cells("A1:D1")
        title_cell = ws.cell(row=1, column=1, value="关键数据指标")
        title_cell.font = Font(name="微软雅黑", size=14, bold=True, color="1F4E79")

        headers = ["指标名称", "数值", "单位", "数据来源"]
        col_widths = [25, 15, 12, 40]
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")

        for c_idx, (header, width) in enumerate(zip(headers, col_widths, strict=True), 1):
            cell = ws.cell(row=3, column=c_idx, value=header)
            cell.font = _HEADER_FONT
            cell.fill = header_fill
            cell.border = _THIN_BORDER
            ws.column_dimensions[get_column_letter(c_idx)].width = width

        for i, metric in enumerate(metrics):
            row = 4 + i
            values = [
                metric.get("metric_name", ""),
                metric.get("metric_value", ""),
                metric.get("metric_unit", ""),
                metric.get("metric_source", ""),
            ]
            for c_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=c_idx, value=value)
                cell.font = _BODY_FONT
                cell.border = _THIN_BORDER

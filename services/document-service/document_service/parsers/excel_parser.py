"""Excel/CSV parser using openpyxl and pandas (M2-13).

Extracts structured tables with column names and sheet names.
"""

from __future__ import annotations

import io
import logging
from typing import Any

from document_service.models import PageContent, ParsedContent, SectionInfo
from document_service.parsers.base import BaseParser

logger = logging.getLogger(__name__)


class ExcelParser(BaseParser):
    """Parse .xlsx/.xls/.csv files using openpyxl, falling back to pandas."""

    def supported_format(self) -> str:
        return "xlsx"

    def parse(self, file_data: bytes, filename: str) -> ParsedContent:
        all_tables: list[dict[str, Any]] = []
        full_text_parts: list[str] = []
        sections: list[SectionInfo] = []

        if filename.lower().endswith(".csv"):
            return self._parse_csv(file_data, filename)

        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(file_data), data_only=True, read_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = []
                for row in ws.iter_rows(values_only=True):
                    rows.append([str(cell) if cell is not None else "" for cell in row])

                if rows:
                    all_tables.append({
                        "sheet": sheet_name,
                        "rows": rows,
                        "columns": list(rows[0]) if rows else [],
                    })

                    # Build text representation
                    full_text_parts.append(f"## Sheet: {sheet_name}")
                    for row in rows:
                        full_text_parts.append(" | ".join(str(c) for c in row))

                    sections.append(SectionInfo(
                        title=sheet_name,
                        level=2,
                        page_start=0,
                        start_char=0,
                    ))

            wb.close()
        except Exception:
            # Fallback to pandas
            return self._parse_pandas_fallback(file_data, filename)

        full_text = "\n".join(full_text_parts)

        return ParsedContent(
            full_text=full_text,
            pages=[PageContent(page_number=1, text=full_text, has_text_layer=True)],
            tables=all_tables,
            sections=sections,
            metadata_hints=self.extract_metadata_hints(file_data, filename),
            needs_ocr=False,
        )

    def _parse_csv(self, file_data: bytes, filename: str) -> ParsedContent:
        """Parse CSV files."""
        import csv as csv_module
        try:
            text = file_data.decode("utf-8")
        except UnicodeDecodeError:
            text = file_data.decode("latin-1", errors="replace")

        reader = csv_module.reader(io.StringIO(text))
        rows = [list(row) for row in reader]

        full_text_parts = []
        for row in rows:
            full_text_parts.append(" | ".join(str(c) for c in row))

        return ParsedContent(
            full_text="\n".join(full_text_parts),
            pages=[PageContent(page_number=1, text="\n".join(full_text_parts), has_text_layer=True)],
            tables=[{"sheet": filename, "rows": rows, "columns": rows[0] if rows else []}],
            sections=[SectionInfo(title=filename, level=1)],
            metadata_hints={"title": filename},
            needs_ocr=False,
        )

    def _parse_pandas_fallback(self, file_data: bytes, filename: str) -> ParsedContent:
        """Fallback parsing using pandas."""
        try:
            import pandas as pd

            if filename.lower().endswith(".csv"):
                df = pd.read_csv(io.BytesIO(file_data))
            else:
                df = pd.read_excel(io.BytesIO(file_data))

            rows = [list(df.columns)] + df.values.tolist()
            rows_cleaned = [[str(c) if c is not None else "" for c in row] for row in rows]

            full_text = "\n".join(" | ".join(str(c) for c in row) for row in rows_cleaned)

            return ParsedContent(
                full_text=full_text,
                pages=[PageContent(page_number=1, text=full_text, has_text_layer=True)],
                tables=[{"sheet": filename, "rows": rows_cleaned, "columns": list(df.columns)}],
                sections=[SectionInfo(title=filename, level=1)],
                metadata_hints={"title": filename},
                needs_ocr=False,
            )
        except ImportError:
            # If pandas also unavailable, return empty
            logger.warning("pandas not available for Excel/CSV fallback")
            return ParsedContent(
                full_text=file_data.decode("utf-8", errors="replace"),
                pages=[PageContent(page_number=1, text=file_data.decode("utf-8", errors="replace"))],
                needs_ocr=False,
            )

    def extract_metadata_hints(self, file_data: bytes, filename: str) -> dict[str, Any]:
        return {"title": filename}
